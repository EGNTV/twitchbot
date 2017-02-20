
import abc
from .abstract.MessageCommand import MessageCommand
from openpyxl import load_workbook
from .abstract.Command import Command
import cfg


class ScoreCommand(MessageCommand):

    def isAllowed(self, username):
        return True

    def getMessage(self, username):
        print(cfg.APP_PATH);
        wb = load_workbook(cfg.APP_PATH+'/files/info.xlsx')
        ws = wb.active
        team1 = {
            "name": ws['B2'].value,
            "score": ws['C2'].value
        }
        team2 = {
            "name": ws['B3'].value,
            "score": ws['C3'].value
        }
        msg = "{} à {} pour {}"
        if team1["score"] > team2["score"]:
            return msg.format(team1["score"],team2["score"],team1["name"])
        else:
            return msg.format(team2["score"],team1["score"],team2["name"])